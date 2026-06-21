import sys
import os
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, Toplevel
from openpyxl import load_workbook

# --- Functions ---


def select_file(entry_widget):
    filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if filename:
        entry_widget.delete(0, tk.END)
        entry_widget.insert(0, filename)

def read_all_values(filepath):
    try:
        wb = load_workbook(filepath, data_only=True)
        values = set()
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None:
                        values.add(str(cell))
        return values
    except Exception as e:
        raise RuntimeError(f"Error reading {filepath}: {e}")

def compare_excels():
    file1 = entry1.get()
    file2 = entry2.get()

    if not file1 or not file2:
        messagebox.showwarning("Missing file", "Please select both Excel files.")
        return

    try:
        values1 = read_all_values(file1)
        values2 = read_all_values(file2)
        matches = values1.intersection(values2)

        output_box.delete(1.0, tk.END)
        if matches:
            output_box.insert(tk.END, f"✅ Found {len(matches)} matching values:\n\n")
            for m in sorted(matches):
                output_box.insert(tk.END, f"- {m}\n")
        else:
            output_box.insert(tk.END, "❌ No matching values found.")
    except Exception as e:
        messagebox.showerror("Error", str(e))

def show_about():
    about_window = Toplevel(root)
    about_window.title("Help / About")
    about_window.geometry("500x350")

    text = (
        "📘 Excel Match Finder - Help / About\n\n"
        "This application allows you to compare two Excel files and find matching cell values.\n"
        "It is designed for quick and easy comparison of spreadsheet contents without needing Excel installed.\n\n"

        "🔧 How It Works:\n"
        "- The program reads all sheets and cells from both files.\n"
        "- Every non-empty cell is collected and compared as text.\n"
        "- Matching values (regardless of their position or sheet) are displayed in a list.\n\n"

        "✅ Steps to Use:\n"
        "1. Click the first 'Browse' button to select the first Excel file.\n"
        "2. Click the second 'Browse' button to select the second Excel file.\n"
        "3. Click 'Compare Files' to begin matching.\n"
        "4. If matches are found, they will be listed in the output box below.\n"
        "5. If no matches are found, a message will notify you.\n\n"

        "⚠️ Important Notes:\n"
        "- Only .xlsx (Excel Open XML) files are supported.\n"
        "- Older .xls files are NOT supported and must be converted to .xlsx before use.\n"
        "- Files saved in LibreOffice or Google Sheets should be exported as .xlsx.\n"
        "- Cell values are treated as plain text (e.g., 123 and '123' are considered the same).\n"
        "- Empty cells and formulas without results are ignored.\n\n"

        "📎 Tip:\n"
        "You can open your .xls file in Excel or LibreOffice and use 'Save As' to export it as .xlsx format.\n\n"

        "👤 Credits:\n"
        "- Developed by: Yunus Sorgunçay\n"
        "- Contact: sorguncay21@itu.edu.tr\n"
        "- Written in: Python (using openpyxl and Tkinter)\n\n"
        # for readers: I omitted the icon-related code before uploading to Github.
        "Version 0.2 - 02/08/2025\n\n"
    )

    help_text = scrolledtext.ScrolledText(about_window, wrap=tk.WORD)
    help_text.insert(tk.END, text)
    help_text.config(state=tk.DISABLED)
    help_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

if __name__ == "__main__":
    # --- GUI Setup ---
    root = tk.Tk()
    root.title("Excel Match Finder")
    root.geometry("600x420")

    # File 1 selection
    tk.Label(root, text="Select File 1 (.xlsx):").pack(anchor="w", padx=10)
    entry1 = tk.Entry(root, width=60)
    entry1.pack(padx=10)
    tk.Button(root, text="Browse", command=lambda: select_file(entry1)).pack(pady=2)

    # File 2 selection
    tk.Label(root, text="Select File 2 (.xlsx):").pack(anchor="w", padx=10)
    entry2 = tk.Entry(root, width=60)
    entry2.pack(padx=10)
    tk.Button(root, text="Browse", command=lambda: select_file(entry2)).pack(pady=2)

    # Compare button
    tk.Button(root, text="Compare Files", command=compare_excels, bg="lightblue").pack(pady=10)

    # Output box
    output_box = scrolledtext.ScrolledText(root, wrap=tk.WORD, height=10)
    output_box.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

    # Help/About button
    tk.Button(root, text="Help / About", command=show_about).pack(pady=5)

    root.mainloop()
